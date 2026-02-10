# -*- coding: utf-8 -*-
"""
@Time ： 2026/2/6 14:16
@Auth ： 章豹
@File ：get_payment_time.py
@IDE ：PyCharm
"""

from typing import List, Dict, Any
from openpyxl import Workbook
from libs.read_yaml import ReadConfig
from libs.sql_service import MySQLService
import os




def _load_config_lists() -> tuple[list, list]:
    """加载 settlement_batch 和 logic_contract_code 配置"""
    base_dir = os.path.dirname(os.path.dirname(__file__))  # 假设 util 是子目录
    contract_path = os.path.join(base_dir, "data", "logic_contract_code.yml")
    batch_path = os.path.join(base_dir, "data", "settlement_batch.yml")

    read_obj = ReadConfig(contract_path)
    list2 = read_obj.read_data("logic_contract_code")

    read_obj = ReadConfig(batch_path)
    list1 = read_obj.read_data("settlement_batch_data")

    if len(list1) != len(list2):
        raise ValueError("settlement_batch 和 logic_contract_code 长度不一致！")

    return list1, list2


def _fetch_records(db, query_type: str, my_dict: Dict[str, str]) -> List[Dict[str, str]]:
    """根据类型查询记录，返回统一格式的列表"""
    endpoint_map = {
        "get_payment": "cash_payment_item",
        "get_cargo": "cargo_acceptance_item",
        "get_receipt": "receipt_acceptance_item"
    }

    if query_type not in endpoint_map:
        raise ValueError(f"不支持的查询类型: {query_type}")

    endpoint = endpoint_map[query_type]
    records = []

    for settlement_batch, logic_contract in my_dict.items():
        try:
            data = db.execute_query(endpoint, {
                'settlement_batch_data': settlement_batch,
                'logic_contract_code': logic_contract
            })

            if data and len(data) > 0:
                record = {
                    'settlement_batch_data': settlement_batch,
                    'logic_contract_code': logic_contract,
                    'instruction_no': data[0].get('instruction_no', 'N/A'),
                    'create_time': data[0]['create_time'].strftime('%Y/%m/%d %H:%M:%S')
                }
            else:
                print(f"未找到数据: {settlement_batch} | {logic_contract}")
                record = {
                    'settlement_batch_data': settlement_batch,
                    'logic_contract_code': logic_contract,
                    'instruction_no': "N/A",
                    'create_time': "N/A"
                }
            records.append(record)

        except Exception as e:
            print(f"查询失败 ({settlement_batch}, {logic_contract}): {e}")
            records.append({
                'settlement_batch_data': settlement_batch,
                'logic_contract_code': logic_contract,
                'instruction_no': "ERROR",
                'create_time': "ERROR"
            })

    return records


def _export_to_excel(records: List[Dict], output_file: str):
    """将记录写入 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "明细"

    headers = ['settlement_batch_data', 'logic_contract_code', 'instruction_no', 'create_time']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, record in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=record["settlement_batch_data"])
        ws.cell(row=row_idx, column=2, value=record["logic_contract_code"])
        ws.cell(row=row_idx, column=3, value=record["instruction_no"])
        ws.cell(row=row_idx, column=4, value=record["create_time"])

    wb.save(output_file)
    print(f"已保存 {len(records)} 条记录到 {output_file}")


def get_sys_time(get_type: str, db_config: Dict[str, Any]):
    """
    获取不同类型的指令创建时间并导出 Excel
    :param get_type: get_payment / get_cargo / get_receipt
    :param db_config: 数据库配置
    """
    # 1. 加载配置
    list1, list2 = _load_config_lists()
    my_dict = dict(zip(list1, list2))
    print("配置映射:", my_dict)

    # 2. 查询数据
    with MySQLService(db_config) as db:
        records = _fetch_records(db, get_type, my_dict)

    # 3. 确定输出文件名
    filename_map = {
        "get_payment": "实际付款日期.xlsx",
        "get_cargo": "实际入库日期_现货.xlsx",
        "get_receipt": "实际入库日期_仓单.xlsx"
    }

    if get_type not in filename_map:
        raise ValueError(f"未知类型: {get_type}")

    output_filename = filename_map[get_type]

    # 4. 构建完整输出路径（保存到 util 目录下）
    util_dir = os.path.dirname(__file__)  # 当前脚本所在目录，即 util/
    print(util_dir)
    output_path = os.path.join(util_dir, output_filename)

    # 5. 如果文件已存在，先删除
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
            print(f"已删除旧文件: {output_path}")
        except OSError as e:
            print(f"无法删除旧文件 {output_path}: {e}")

    # 6. 导出新文件
    _export_to_excel(records, output_path)  # 注意：传完整路径！,目前这边直接生成到当前路径，所以没做修改

    print(f"\n{get_type} 共 {len(records)} 条记录")


if __name__ == '__main__':
    db_config = {
        'host': '192.168.100.35',
        'user': 'mgmt',
        'password': 'mgmt',
        'database': 'mgmt_data'
    }
    get_sys_time("get_cargo", db_config)