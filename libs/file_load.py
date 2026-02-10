# -*- coding: utf-8 -*-
"""
@Time ： 2026/2/9 15:39
@Auth ： 章豹
@File ：file_load.py
@IDE ：PyCharm
"""
import openpyxl
from paths_manger import login_data_xlsx


def read_excel(filepath,sheet_name):
    """
    读取excel数据
    :param filepath: 文件路径
    :param sheet_name: 读取文件的sheet名称
    :return:
    """
    wb = openpyxl.load_workbook(filepath)
    sheet_data = wb[sheet_name]
    lines_count = sheet_data.max_row # 获取总行数
    cols_count = sheet_data.max_column # 获取总列数
    data = []
    for i in range(2,lines_count+1):
        line = [] # 用来存储所有行的数据，每行数据都是这个列表的子列表
        for c in range(1,cols_count+1):
            cell_data = sheet_data.cell(i,c).value
            line.append(cell_data)
        data.append(line)
    return  data

# if __name__ == '__main__':
#     print(read_excel(login_data_xlsx, "Sheet1"))